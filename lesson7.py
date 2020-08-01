# -*- coding: utf-8 -*-
"""
@Time ： 2020/7/27 20:14
@Author ： YaMeng
@File ：lesson7.py
@IDE ：PyCharm
@CopyRight：湖南省零檬信息技术有限公司
"""

'''
你们所认为的python自动化工作应该是个什么样子？
1、准备好自动化的测试用例    === done   test_case_api.xlsx
2、使用python去读取测试用例 === done   read_data()
4、发送请求，得到响应结果    === done   api_func()
5、结果的判断？ 执行结果 vs  预期结果  == 断言
6、得到一个最终结果，回写到测试用例   === done   write_result()
'''

'''
{'case_id': 3, 
'url': 'http://api.lemonban.com/futureloan/member/register', 
'data': '{"pwd":"12345678","type":1}', 
'expected': '{"code":1,"msg":"手机号为空"}'}
'''
'''
比如 a = '10 * 20'
eval(a) 
10 * 20
'''

import openpyxl
import requests


# 读取测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 取出sheet里最大的行数
    # print(max_row)
    case_list = []
    for i in range(2, max_row + 1, 1):  # 取值是取左不取右，左闭右开
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,  # 取出url
            data=sheet.cell(row=i, column=6).value,  # 取出请求体
            expected=sheet.cell(row=i, column=7).value  # 取出预期结果
        )
        case_list.append(dict1)  # dict1里面是一条一条的测试用例， --->装到列表里面  这个列表就存放了所有的测试用例
    # print(case_list)
    return case_list


# 发送请求
def api_func(url, data):
    header_login = {'X-Lemonban-Media-Type': 'lemonban.v2',
                    'Content-Type': 'application/json'}
    res1 = requests.post(url=url, json=data, headers=header_login)
    # print(res1.json())
    response = res1.json()
    return response


# 写入测试结果
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)


# 封装成一个执行函数并断言
def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)  # 读取excel里的测试用例
    # print(cases)
    for case in cases:  # 循环取出测试用例
        case_id = case.get('case_id')  # 字典取值
        url = case['url']  # 取url通过key的方法
        data = case.get('data')  # 通过excel取出的值是str
        data = eval(data)  # eval() 作用：运行被字符串包括的python表达式
        expected = case.get('expected')
        expected = eval(expected)
        expected_msg = expected.get('msg')  # 取出预期结果里的msg信息
        real_result = api_func(url=url, data=data)  # == 调用了发送请求的函数 并且 传入参数
        real_msg = real_result.get('msg')  # 取出实际执行结果里的msg信息
        print('预期结果为：{}'.format(expected_msg))
        print('实际结果为：{}'.format(real_msg))
        if real_msg == expected_msg:
            print('第{}条用例通过！'.format(case_id))
            final_res = 'pass'
        else:
            print('第{}条用例未通过！'.format(case_id))
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, case_id + 1, 8, final_res)


execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')



